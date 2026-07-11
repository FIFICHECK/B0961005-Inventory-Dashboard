#!/usr/bin/env python3
import openpyxl, json, re, shutil, os
from collections import defaultdict
from datetime import datetime

# === 1. Copy new file ===
src = "/home/snkwok/.hermes/profiles/hermes1/cache/documents/doc_99cb0d731f1a_ECOM-MMSNG_DAILY_ORDER_B0961005_20260711235959.xlsx"
dst_dir = "/home/snkwok/B0961005-Inventory-Dashboard/reports/order_reports/"
dst = os.path.join(dst_dir, "ECOM-MMSNG_DAILY_ORDER_B0961005_20260711235959.xlsx")
shutil.copy2(src, dst)
print("Copied to", dst)

# === 2. Parse new file ===
wb = openpyxl.load_workbook(dst)
ws = wb.active
total_gmv = round(ws.cell(2, 6).value, 2)
order_count = ws.max_row - 4

hourly = defaultdict(float)
for r in range(6, ws.max_row + 1):
    ot = ws.cell(r, 8).value
    tv = ws.cell(r, 27).value
    if ot and tv:
        h = int(str(ot).split(':')[0])
        hourly[h] += float(tv)
hourly_list = [round(hourly.get(h, 0), 2) for h in range(24)]

print(f"GMV: ${total_gmv:.2f}, Orders: {order_count}")

# === 3. Update order_data.json ===
with open('/home/snkwok/B0961005-Inventory-Dashboard/data/order_data.json', 'r') as f:
    order_data = json.load(f)

order_data['gmv_daily']['2026-07-11'] = total_gmv
order_data['gmv_hourly']['2026-07-11'] = {str(h): hourly_list[h] for h in range(24)}
if 'order_count_daily' not in order_data:
    order_data['order_count_daily'] = {}
order_data['order_count_daily']['2026-07-11'] = order_count

with open('/home/snkwok/B0961005-Inventory-Dashboard/data/order_data.json', 'w') as f:
    json.dump(order_data, f, indent=2)
print("order_data.json updated")

# === 4. Update manifest ===
with open('/home/snkwok/B0961005-Inventory-Dashboard/data/order_reports_manifest.json', 'r') as f:
    manifest = json.load(f)

for entry in manifest:
    if entry['date'] == '7月11日':
        old_entry = entry.copy()
        entry['gmv'] = f"${total_gmv:,.2f}"
        entry['timestamp'] = '23:59'
        entry['filename'] = 'ECOM-MMSNG_DAILY_ORDER_B0961005_20260711235959.xlsx'
        break

with open('/home/snkwok/B0961005-Inventory-Dashboard/data/order_reports_manifest.json', 'w') as f:
    json.dump(manifest, f, indent=2)
print("manifest updated")

# === 5. Regenerate sales_trend_data.js ===
dates = sorted(order_data['gmv_daily'].keys())
gmv_vals = [round(order_data['gmv_daily'][d], 2) for d in dates]
running = 0
cumulative = []
for v in gmv_vals:
    running += v
    cumulative.append(round(running, 2))

js_content = '// Auto-generated sales trend data\n'
js_content += '// Generated: 2026-07-12 01:00\n'
js_content += '// Source: order_data.json (includes Jul11 23:59 final)\n\n'
js_content += 'const salesTrendData = {\n'
js_content += '  "gmv_by_date": {\n'
js_content += '    "labels": ' + json.dumps(dates) + ',\n'
js_content += '    "values": ' + json.dumps(gmv_vals) + ',\n'
js_content += '    "cumulative": ' + json.dumps(cumulative) + '\n'
js_content += '  }\n'
js_content += '};\n'

with open('/home/snkwok/B0961005-Inventory-Dashboard/data/sales_trend_data.js', 'w') as f:
    f.write(js_content)
print(f"sales_trend_data.js regenerated (total: ${cumulative[-1]:,.2f})")

# === 6. Update inline salesTrendData in index.html ===
with open('/home/snkwok/B0961005-Inventory-Dashboard/index.html', 'rb') as f:
    html = f.read().decode('utf-8')

pattern = r'const salesTrendData\s*=\s*(\{.*?\});'
match = re.search(pattern, html, re.DOTALL)
idata = json.loads(match.group(1))

# Get old values
jul11_idx = idata['gmv_by_date']['labels'].index('2026-07-11')
old_gmv = idata['gmv_by_date']['data'][jul11_idx]
old_orders_partial = idata['orders_by_date']['data'][jul11_idx]
delta = round(total_gmv - old_gmv, 2)
jul11_order_delta = order_count - old_orders_partial

print(f"Old Jul11 GMV: {old_gmv}, New: {total_gmv}, Delta: {delta}")
print(f"Old Jul11 Orders: {old_orders_partial}, New: {order_count}, Delta: {jul11_order_delta}")

# Update gmv_by_date
idata['gmv_by_date']['data'][jul11_idx] = total_gmv

# Update gmv_by_day_of_week
for i, l in enumerate(idata['gmv_by_day_of_week']['labels']):
    if '2026-07-11' in l:
        idata['gmv_by_day_of_week']['data'][i] = total_gmv
        break

# Update gmv_by_month (July)
jul_idx = idata['gmv_by_month']['labels'].index('2026-07')
idata['gmv_by_month']['data'][jul_idx] = round(idata['gmv_by_month']['data'][jul_idx] + delta, 2)

# Update orders
idata['orders_by_date']['data'][jul11_idx] = order_count
jul_ord_idx = idata['orders_by_month']['labels'].index('2026-07')
idata['orders_by_month']['data'][jul_ord_idx] += jul11_order_delta

# Update hourly data
idata['gmv_by_date_hour']['data']['2026-07-11'] = hourly_list

# Recalculate gmv_by_hour
all_hours = [0.0] * 24
for date_key, hvals in idata['gmv_by_date_hour']['data'].items():
    for h in range(min(len(hvals), 24)):
        all_hours[h] += hvals[h]
idata['gmv_by_hour']['data'] = [round(v, 2) for v in all_hours]

# Update summary
idata['summary']['total_gmv'] = round(idata['summary']['total_gmv'] + delta, 2)
idata['summary']['total_orders'] += jul11_order_delta
idata['summary']['avg_order_value'] = round(idata['summary']['total_gmv'] / idata['summary']['total_orders'], 2)
idata['summary']['this_month']['gmv'] = round(idata['summary']['this_month']['gmv'] + delta, 2)
idata['summary']['this_month']['orders'] += jul11_order_delta
idata['summary']['this_month']['avg'] = round(idata['summary']['this_month']['gmv'] / idata['summary']['this_month']['orders'], 2)

new_json = json.dumps(idata, ensure_ascii=False, separators=(',', ':'))
html = html.replace(match.group(0), f'const salesTrendData = {new_json};')
print("Inline salesTrendData updated")

# === 7. Update Report tab in index.html ===
# Replace the July 11 row (18:00 → 23:59, GMV, download link)
old_row = '<strong>2026-07-11</strong></td>'
old_row += '\n                    <td class="align-middle text-muted text-center">18:00</td>'
old_row += '\n                    <td class="align-middle text-end">$134,418.82</td>'
old_row += '\n                    <td class="align-middle">'
old_row += '\n                        <a href="reports/order_reports/ECOM-MMSNG_DAILY_ORDER_B0961005_20260711180000.xlsx" download'

new_row = '<strong>2026-07-11</strong></td>'
new_row += '\n                    <td class="align-middle text-muted text-center">23:59</td>'
gmv_str = f"${total_gmv:,.2f}"
new_row += f'\n                    <td class="align-middle text-end">{gmv_str}</td>'
new_row += '\n                    <td class="align-middle">'
new_row += '\n                        <a href="reports/order_reports/ECOM-MMSNG_DAILY_ORDER_B0961005_20260711235959.xlsx" download'

html = html.replace(old_row, new_row)

# Update header
html = html.replace(
    'Order Report: 2026-07-10 (23:59:59)',
    'Order Report: 2026-07-11 (23:59:59)'
)

with open('/home/snkwok/B0961005-Inventory-Dashboard/index.html', 'wb') as f:
    f.write(html.encode('utf-8'))
print("index.html Report tab & header updated")

# === Verify ===
print(f"\n=== VERIFICATION ===")
print(f"Jul11 GMV: ${total_gmv:,.2f}")
print(f"Jul11 Orders: {order_count}")
print(f"July monthly GMV: ${idata['summary']['this_month']['gmv']:,.2f}")
print(f"Total GMV: ${idata['summary']['total_gmv']:,.2f}")
print(f"Total Orders: {idata['summary']['total_orders']}")
